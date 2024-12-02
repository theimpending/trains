from math import floor
import openpyxl

#import data from excel sheet/make sure to replace the path file with the path of the excel sheet
user = input("Enter the path of the excel sheet: ")
try:
    wb = openpyxl.load_workbook(rf"{user}")
except openpyxl.utils.exceptions.InvalidFileException:
    print("Error: The file you entered is not an excel file or not in the right format")
    exit(2)
timesheet = wb['TimeSheet']
options = wb['Options']
# Station/total people in and out by their respective time
station = {}
index = 2
for col in timesheet['A'][1:]:
    if col.value is not None:
        station[col.value] = []
        for row in timesheet[index][1:]:
            if row.value is not None:
                station[col.value].append(row.value)
        index += 1

        

time = [cell.value for cell in timesheet[1][1:] if cell.value is not None]



# the distance between each station in miles
station_dist = [cell.value for cell in options['B'][1:] if cell.value is not None]

# At what percentange do you want there to be a rapid station
threshold_for_rapid = options['H1'].value

# total amount of usuable trains
trains = options['H2'].value

# train speed in mph
train_speed = options['H4'].value

# train capacity
train_cap = options['H3'].value

# stored data
rapid_sta = []
total_station = len(station)
rapid_sta_dist = []
time_and_train_min = {}
total_pass_per_time = []
total_pass = 0
dispatch_time = []
schedule = {}
station_list = list(station.keys())
pass_num_per_station = []
train_used = 0


# created functions
# to find distance between each rapid
def dist_find(a, b):
    index_a = station_list.index(a)
    index_b = station_list.index(b)
    distance = 0
    if index_a < index_b:
        while index_a < index_b:
            distance += station_dist[index_a]
            index_a += 1
    else:
        while index_a > index_b:
            distance += station_dist[index_a - 1]
            index_a -= 1
    return distance

# returns the times dispatched in time format
def hourly_dispatch(time, disp_train):
    dispatch = 60 / disp_train
    hour, minute = time.split(":")
    minute = int(minute)
    hourly_schedule = []
    for x in range(disp_train):
        hourly_schedule.append(f'{hour}:{int(minute):02}')
        minute += dispatch
    return hourly_schedule

def train_schedule(time, distance, train_speed):
    time = time.split(":")
    hour = int(time[0])
    minute = int(time[1])
    time_in_minutes = (hour * 60) + minute
    time_to_travel = distance / train_speed
    time_in_minutes += (time_to_travel * 60)
    hour = time_in_minutes // 60
    minute = time_in_minutes % 60
    return f'{int(hour)}:{int(minute):02}'

    


# check to make sure the data inputed is correct

#this make sure that the amount of people in each station is the same as the amount of time
for x in station:
    number_of_time_in_station = len(station[x])
    if number_of_time_in_station != len(time):
        print("Error: There is a mismatch between the amount of time and the amount of people inputed")
        exit(1)
    else:
        continue
#this adds the average amount of people in each station to the list and add total passengers
for x in station:
    pass_num = 0
    for person in station[x]:
        pass_num += person
    pass_num_per_station.append(pass_num)
    total_pass += pass_num
average_percentage_per_station = [round(x / total_pass, 2) for x in pass_num_per_station]

#this adds the total amount of people in each time to the list
for x in time:
    pass_num = 0
    for y in station:
        pass_num += station[y][time.index(x)]
    total_pass_per_time.append(pass_num)

try:
    float(threshold_for_rapid)
except ValueError:
    print("Error: The threshold should be a number")
    exit(3)
#------------------------------------------------------------
# choosing the rapid station
for x in range(total_station):
    if average_percentage_per_station[x] >= threshold_for_rapid:
        rapid_sta.append(station_list[x])


# find distance between each rapid station
for x in range(len(rapid_sta) - 1):
    rapid_sta_dist.append(dist_find(rapid_sta[x], rapid_sta[x+1]))



# train schedule generator ---------------------------------
# to find out how many trains to dispatch per hour
for x in range(len(time)):
    train_dispatch = total_pass_per_time[x] / total_pass
    train_dispatch = float(f'{train_dispatch:.02f}')
    train_per_hour = floor(train_dispatch * trains)
    time_and_train_min[time[x]] = train_per_hour
    train_used += train_per_hour


# train dispatch times
time_and_train_min_list = list(time_and_train_min.keys())
for x in range(len(time_and_train_min)):
    if time_and_train_min[time_and_train_min_list[x]] == 0:
        continue
    hourly_time = hourly_dispatch(time_and_train_min_list[x],time_and_train_min[time_and_train_min_list[x]])
    dispatch_time += hourly_time

# schedule for each station
for x in rapid_sta:
    schedule[x] = []
    if rapid_sta.index(x) == 0:
        schedule[x] = dispatch_time
    else:
        for y in schedule[rapid_sta[rapid_sta.index(x) - 1]]:
            updated = train_schedule(y, rapid_sta_dist[rapid_sta.index(x) - 1], train_speed)
            schedule[x].append(updated)


#------------------------------------------------------------
# save the data to an excel sheet
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Schedule"
sheet2 = wb.create_sheet(title="Data")
# Write headers to the Schedule sheet
sheet['A1'] = 'Rapid Station/Times'

# Write data to the Schedule sheet
index = 2
for x in rapid_sta:
    sheet.cell(row=index, column=1, value=x)
    for col, y in enumerate(schedule[x], start=2):
        sheet.cell(row=index, column=col, value=y)
    index += 1
# Write data to the worksheet
sheet2['A1'] = "Total Passengers"
sheet2['B1'] = total_pass

sheet2['A2'] = "Total Trains Used"
sheet2['B2'] = train_used

sheet2['D2'] = "Trains Needed"
sheet2['E2'] = floor(total_pass / train_cap)

sheet2['A3'] = "Station Names"
for col, value in enumerate(station_list, start=2):
    sheet2.cell(row=3, column=col, value=value)

sheet2['A4'] = "Total Passenger per Station"
for col, value in enumerate(pass_num_per_station, start=2):
    sheet2.cell(row=4, column=col, value=value)

sheet2['A5'] = "Average Percentage per Station"
for col, value in enumerate(average_percentage_per_station, start=2):
    sheet2.cell(row=5, column=col, value=value)

sheet2['A6'] = "Rapid Station Names"
for col, value in enumerate(rapid_sta, start=2):
    sheet2.cell(row=6, column=col, value=value)

file_save = "train_schedule.xlsx"
wb.save(file_save)
print(f"Data has been saved to {file_save}")

